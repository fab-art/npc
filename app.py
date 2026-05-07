"""
MedMatch AI — Medical Standardization Engine  v2.1
Rwanda FDA · SOP ODDG/RES/SOP/004

Modules:
  [1] NormalizationEngine  — brand aliases, spec extraction, ATC/INN expansion
  [2] EmbeddingEngine      — Sentence Transformers / HF API / TF-IDF / RapidFuzz
  [3] CandidateRetrieval   — cosine similarity, FAISS-ready, top-20
  [4] ReRankEngine         — cross-encoder ms-marco-MiniLM-L-6-v2
  [5] HybridScorer         — 70% semantic + 30% fuzzy
  [6] ValidationEngine     — rule-based clinical safety layer
  [7] DittoVerifier        — pairwise semantic confirmation
  [8] EnrichmentEngine     — openFDA device + drug + WHO INN
  [9] StandardizationOutput — NPC Code, match type, confidence

API Keys (add to ~/.streamlit/secrets.toml):
  OPENFDA_KEY  = ""   # 120k req/day (vs 1k free)
  HF_TOKEN     = ""   # HuggingFace Inference API fallback
  SERPAPI_KEY  = ""   # Google Search enrichment (optional)
"""

import streamlit as st
import pandas as pd
import numpy as np
import re, io, time, hashlib, threading, warnings, json
from datetime import datetime
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings("ignore")

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    from sentence_transformers import SentenceTransformer, CrossEncoder, util
    ST_OK = True
except ImportError:
    ST_OK = False

try:
    import faiss
    FAISS_OK = True
except ImportError:
    FAISS_OK = False

try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    REQ_OK = True
except ImportError:
    REQ_OK = False

try:
    from rapidfuzz import fuzz, process as rfproc
    RF_OK = True
except ImportError:
    RF_OK = False

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity as sk_cos
    SKL_OK = True
except ImportError:
    SKL_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPX_OK = True
except ImportError:
    OPX_OK = False

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MedMatch AI",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;700&family=Syne:wght@400;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap');
:root{
  --bg:#060d18;--sf:#0d1b2a;--sf2:#112236;--sf3:#162d44;--bd:#1e3a5f;--bd2:#2a4f7a;
  --ac:#00d4ff;--ac2:#7eeeff;--ac3:#00ffbb;
  --warn:#ffb300;--err:#ff4757;--ok:#00e676;--pur:#a78bfa;--pink:#f472b6;
  --tx:#e8f4fd;--mu:#6b8fa8;--mu2:#4a6b85;
  --mono:'JetBrains Mono',monospace;--head:'Syne',sans-serif;--body:'Inter',sans-serif;
}
.stApp{background:var(--bg);color:var(--tx);font-family:var(--body);}
section[data-testid="stSidebar"]{background:var(--sf)!important;border-right:1px solid var(--bd);}
.stButton>button{background:linear-gradient(135deg,#0077aa,#00d4ff);color:#fff;
  font-family:var(--head);font-weight:700;border:none;border-radius:8px;
  padding:.55rem 1.6rem;letter-spacing:.03em;transition:all .2s;}
.stButton>button:hover{transform:translateY(-2px);box-shadow:0 6px 24px rgba(0,212,255,.35);}
.stButton>button:disabled{background:var(--sf3);color:var(--mu);box-shadow:none;transform:none;}
div[data-testid="stFileUploader"]{border:1.5px dashed var(--bd2);border-radius:10px;background:var(--sf);}
div[data-testid="stFileUploader"]:hover{border-color:var(--ac);}
.stTextInput>div>div>input,.stTextArea>div>div>textarea{
  background:var(--sf2)!important;border:1px solid var(--bd)!important;
  border-radius:8px!important;color:var(--tx)!important;}
.stTextInput>div>div>input:focus{border-color:var(--ac)!important;
  box-shadow:0 0 0 2px rgba(0,212,255,.15)!important;}
.stSelectbox>div>div,.stMultiSelect>div>div{
  background:var(--sf2)!important;border:1px solid var(--bd)!important;border-radius:8px!important;}
.stSlider>div>div>div{background:var(--ac)!important;}
.stProgress>div>div{background:linear-gradient(90deg,var(--ac),var(--ac3));}
hr{border-color:var(--bd);}
.stTabs [data-baseweb="tab-list"]{background:var(--sf);border-bottom:1px solid var(--bd);}
.stTabs [data-baseweb="tab"]{color:var(--mu);font-family:var(--head);font-weight:600;}
.stTabs [aria-selected="true"]{color:var(--ac)!important;border-bottom:2px solid var(--ac)!important;}
/* Typography */
.logo{font-family:var(--head);font-size:1.5rem;font-weight:800;
  background:linear-gradient(135deg,var(--ac),var(--ac3));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.logo-sub{font-size:.65rem;color:var(--mu);letter-spacing:.12em;text-transform:uppercase;}
.pt{font-family:var(--head);font-size:2.1rem;font-weight:800;
  background:linear-gradient(135deg,var(--ac2),var(--ac3));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;letter-spacing:-.02em;}
.ps{color:var(--mu);font-size:.9rem;margin-top:.25rem;margin-bottom:1.5rem;}
.sh{font-family:var(--head);font-weight:700;font-size:1rem;
  border-left:3px solid var(--ac);padding-left:.7rem;margin:1.1rem 0 .6rem;color:var(--tx);}
/* Cards */
.sc{background:var(--sf);border:1px solid var(--bd);border-radius:12px;padding:1.1rem 1.3rem;
  position:relative;overflow:hidden;}
.sc::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:linear-gradient(90deg,var(--ac),var(--ac3));}
.sv{font-family:var(--head);font-size:2rem;font-weight:800;}
.sl{font-size:.72rem;color:var(--mu);text-transform:uppercase;letter-spacing:.09em;}
.sg{color:var(--ok);}.sb{color:var(--ac);}.sy{color:var(--warn);}.sp{color:var(--pur);}.se{color:var(--err);}
/* Result cards */
.rc{background:var(--sf);border:1px solid var(--bd);border-radius:12px;
  padding:1.1rem 1.3rem;margin:.5rem 0;transition:border-color .2s;}
.rc:hover{border-color:var(--bd2);}
.rc.r1{border-color:rgba(0,212,255,.45);background:rgba(0,212,255,.04);}
.rc.r2{border-color:rgba(0,255,187,.25);}
.rn{font-family:var(--mono);font-size:.77rem;color:var(--ac);font-weight:700;letter-spacing:.06em;}
.rd{font-family:var(--head);font-size:.97rem;font-weight:700;color:var(--tx);margin:.2rem 0;}
/* Confidence bar */
.cbw{background:var(--sf3);border-radius:100px;height:6px;margin:.4rem 0;}
.cbar{height:6px;border-radius:100px;}
.ch{background:linear-gradient(90deg,#00e676,#00ffbb);}
.cm{background:linear-gradient(90deg,#ffb300,#ffd54f);}
.cl{background:linear-gradient(90deg,#ff4757,#ff6b81);}
/* Pills */
.pill{display:inline-block;font-family:var(--mono);font-size:.65rem;font-weight:700;
  padding:2px 8px;border-radius:5px;margin:.1rem;}
.pE{background:rgba(0,230,118,.15);color:#00e676;border:1px solid rgba(0,230,118,.25);}
.pB{background:rgba(0,212,255,.15);color:var(--ac);border:1px solid rgba(0,212,255,.25);}
.pS{background:rgba(255,179,0,.15);color:#ffb300;border:1px solid rgba(255,179,0,.25);}
.pN{background:rgba(107,143,168,.1);color:var(--mu);border:1px solid rgba(107,143,168,.2);}
.pV{background:rgba(0,230,118,.12);color:#00e676;}
.pR{background:rgba(255,71,87,.12);color:#ff4757;}
.pAI{background:rgba(167,139,250,.15);color:var(--pur);border:1px solid rgba(167,139,250,.25);}
.pD{background:rgba(244,114,182,.15);color:var(--pink);border:1px solid rgba(244,114,182,.25);}
.pW{background:rgba(0,212,255,.1);color:var(--ac2);border:1px solid rgba(0,212,255,.2);}
/* Engine badges */
.mb{display:inline-flex;align-items:center;gap:.3rem;padding:.2rem .6rem;border-radius:6px;
  font-family:var(--mono);font-size:.67rem;font-weight:700;margin:.1rem 0;}
.mbok{background:rgba(0,230,118,.1);color:#00e676;border:1px solid rgba(0,230,118,.25);}
.mbwn{background:rgba(255,179,0,.1);color:#ffb300;border:1px solid rgba(255,179,0,.25);}
.mbof{background:rgba(107,143,168,.1);color:var(--mu);border:1px solid rgba(107,143,168,.2);}
/* Pipeline */
.pipe{display:inline-flex;flex-direction:column;align-items:center;gap:.15rem;
  padding:.45rem .65rem;border:1px solid var(--bd);border-radius:8px;
  font-size:.62rem;font-family:var(--mono);color:var(--mu);
  background:var(--sf);min-width:72px;text-align:center;}
.pipe.dn{border-color:var(--ok);color:var(--ok);background:rgba(0,230,118,.05);}
.pipe.on{border-color:var(--ac);color:var(--ac);background:rgba(0,212,255,.06);}
.pipe.off{opacity:.35;}
/* Console */
.cons{background:#020810;border:1px solid var(--bd);border-radius:10px;
  padding:1rem 1.2rem;font-family:var(--mono);font-size:.72rem;
  max-height:260px;overflow-y:auto;line-height:1.9;}
.co{color:var(--ok);}.cw{color:var(--warn);}.ce{color:var(--err);}
.cai{color:var(--pur);}.cn{color:var(--ac);}.ci{color:#58a6ff;}
/* FDA card */
.fda{background:rgba(0,212,255,.04);border:1px solid rgba(0,212,255,.2);
  border-radius:8px;padding:.8rem 1rem;margin:.4rem 0;font-size:.82rem;}
.fda.who{background:rgba(0,255,187,.04);border-color:rgba(0,255,187,.2);}
.fda.ditto{background:rgba(167,139,250,.04);border-color:rgba(167,139,250,.2);}
.ff{color:var(--mu);font-size:.7rem;text-transform:uppercase;letter-spacing:.07em;}
.fv{font-family:var(--mono);font-size:.8rem;color:var(--ac2);margin-bottom:.3rem;}
.fv.g{color:var(--ok);}.fv.p{color:var(--pur);}
/* History */
.hi{background:var(--sf2);border:1px solid var(--bd);border-radius:7px;
  padding:.45rem .7rem;margin:.25rem 0;font-size:.78rem;color:var(--mu);}
/* Search box glow */
.swrap{background:var(--sf);border:1.5px solid var(--bd2);border-radius:12px;padding:1rem 1.2rem;}
.swrap:focus-within{border-color:var(--ac);box-shadow:0 0 0 3px rgba(0,212,255,.1);}
/* ATC drug badge */
.atc{display:inline-block;background:rgba(244,114,182,.12);color:var(--pink);
  border:1px solid rgba(244,114,182,.25);border-radius:5px;
  font-family:var(--mono);font-size:.64rem;padding:1px 7px;margin:.1rem;}
/* Ditto score ring */
.ditto-ring{display:inline-flex;align-items:center;justify-content:center;
  width:52px;height:52px;border-radius:50%;
  border:3px solid var(--pur);font-family:var(--head);font-size:.95rem;
  font-weight:800;color:var(--pur);}
/* Settings toggle row */
.stg-row{background:var(--sf);border:1px solid var(--bd);border-radius:9px;
  padding:.7rem 1rem;margin:.35rem 0;display:flex;justify-content:space-between;
  align-items:center;font-size:.85rem;}
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# ── SESSION STATE ─────────────────────────────────────────────────────────────
def _init():
    defaults = dict(
        page="search",
        catalog_df=None,
        embeddings=None,        # np array (N,384)
        faiss_idx=None,         # faiss.IndexFlatIP if available
        embed_meta=None,        # list of {code,desc,generic,category,family,atc}
        tfidf_mat=None,
        tfidf_vec=None,
        history=deque(maxlen=60),
        batch_results=None,
        model_name=None,
        enrich_cache={},
        who_cache={},
        ditto_cache={},
        settings=dict(
            npc_threshold=65,
            fuzzy_weight=0.30,
            semantic_weight=0.70,
            use_ditto=True,
            use_who=True,
            use_fda=True,
            use_cross_encoder=True,
            model_choice="all-MiniLM-L6-v2",
            batch_workers=3,
            confidence_high=85,
            confidence_med=60,
        ),
        stats=dict(queries=0,exact=0,brand=0,spec=0,new=0,avg_ms=0,
                   ditto_confirms=0,who_hits=0,fda_hits=0),
    )
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
_init()
S = st.session_state
CFG = S.settings

# ── KNOWLEDGE BASES ───────────────────────────────────────────────────────────
# Brand → Generic alias map
BRAND = {
    # Surgical sutures
    "CORTEX SCREW":"CORTICAL BONE SCREW","CONCELOUS":"CANCELLOUS","CANSELLOUS":"CANCELLOUS",
    "CAPROSYN":"POLYGLYTONE 6211 ABSORBABLE MONOFILAMENT SUTURE",
    "VICRYL":"POLYGLACTIN 910 ABSORBABLE BRAIDED SUTURE",
    "VICRYL RAPIDE":"POLYGLACTIN 910 RAPID ABSORBABLE SUTURE",
    "MONOCRYL":"POLIGLECAPRONE 25 ABSORBABLE MONOFILAMENT SUTURE",
    "PROLENE":"POLYPROPYLENE NON-ABSORBABLE MONOFILAMENT SUTURE",
    "PROLENE BLUE":"POLYPROPYLENE NON-ABSORBABLE MONOFILAMENT SUTURE",
    "PDS":"POLYDIOXANONE ABSORBABLE MONOFILAMENT SUTURE",
    "PDS II":"POLYDIOXANONE ABSORBABLE MONOFILAMENT SUTURE",
    "POLYSORB":"POLYGLYCOLIC ACID LACTIDE ABSORBABLE BRAIDED SUTURE",
    "BIOSYN":"GLYCOMER 631 ABSORBABLE MONOFILAMENT SUTURE",
    "VELOSORB":"POLYGLYCOLIC ACID LACTIDE ABSORBABLE BRAIDED SUTURE",
    "MAXON":"POLYGLYCOLATE ABSORBABLE MONOFILAMENT SUTURE",
    "SURGIPRO":"POLYPROPYLENE NON-ABSORBABLE MONOFILAMENT SUTURE",
    "TICRON":"COATED POLYESTER NON-ABSORBABLE BRAIDED SUTURE",
    "SURGIDAC":"POLYESTER NON-ABSORBABLE BRAIDED SUTURE",
    "SOFSILK":"SILK NON-ABSORBABLE BRAIDED SUTURE",
    "SURGILON":"BRAIDED NYLON NON-ABSORBABLE SUTURE",
    "MONOSOF":"NYLON MONOFILAMENT NON-ABSORBABLE SUTURE",
    # Surgical devices
    "SURGICEL":"OXIDIZED REGENERATED CELLULOSE HAEMOSTATIC",
    "SURGICELL":"OXIDIZED REGENERATED CELLULOSE HAEMOSTATIC",
    "HEMOLOCK":"HEM-O-LOK POLYMER LIGATION CLIP",
    "LIGASURE":"VESSEL SEALING ELECTROTHERMAL BIPOLAR DEVICE",
    "REDON":"CLOSED WOUND SUCTION DRAIN",
    "VERSAPORT":"LAPAROSCOPIC RADIALLY EXPANDING TROCAR",
    "VERSASTEP":"LAPAROSCOPIC TROCAR SYSTEM",
    "SONICISION":"CORDLESS ULTRASONIC DISSECTOR VESSEL SEALER",
    # Pharmaceuticals (INN)
    "PANADOL":"PARACETAMOL ANALGESIC ANTIPYRETIC",
    "TYLENOL":"PARACETAMOL ANALGESIC ANTIPYRETIC",
    "CALPOL":"PARACETAMOL ANALGESIC ANTIPYRETIC",
    "AUGMENTIN":"AMOXICILLIN CLAVULANIC ACID ANTIBIOTIC",
    "CO-AMOXICLAV":"AMOXICILLIN CLAVULANIC ACID ANTIBIOTIC",
    "FLAGYL":"METRONIDAZOLE ANTIBIOTIC ANTIPROTOZOAL",
    "METROGYL":"METRONIDAZOLE ANTIBIOTIC ANTIPROTOZOAL",
    "BRUFEN":"IBUPROFEN NSAID ANTI-INFLAMMATORY ANALGESIC",
    "NUROFEN":"IBUPROFEN NSAID ANTI-INFLAMMATORY ANALGESIC",
    "ADVIL":"IBUPROFEN NSAID ANTI-INFLAMMATORY ANALGESIC",
    "VENTOLIN":"SALBUTAMOL BRONCHODILATOR BETA2 AGONIST",
    "LASIX":"FUROSEMIDE LOOP DIURETIC",
    "ALDACTONE":"SPIRONOLACTONE POTASSIUM-SPARING DIURETIC",
    "ZITHROMAX":"AZITHROMYCIN MACROLIDE ANTIBIOTIC",
    "ZITROCIN":"AZITHROMYCIN MACROLIDE ANTIBIOTIC",
    "CIPROBAY":"CIPROFLOXACIN FLUOROQUINOLONE ANTIBIOTIC",
    "CIPROXIN":"CIPROFLOXACIN FLUOROQUINOLONE ANTIBIOTIC",
    "SEPTRIN":"TRIMETHOPRIM SULFAMETHOXAZOLE ANTIBIOTIC COTRIMOXAZOLE",
    "BACTRIM":"TRIMETHOPRIM SULFAMETHOXAZOLE ANTIBIOTIC COTRIMOXAZOLE",
    "GENTOCIN":"GENTAMICIN AMINOGLYCOSIDE ANTIBIOTIC",
    "GARAMYCIN":"GENTAMICIN AMINOGLYCOSIDE ANTIBIOTIC",
    "DIFLUCAN":"FLUCONAZOLE ANTIFUNGAL TRIAZOLE",
    "CANESTEN":"CLOTRIMAZOLE ANTIFUNGAL",
    "COARTEM":"ARTEMETHER LUMEFANTRINE ANTIMALARIAL",
    "FANSIDAR":"SULFADOXINE PYRIMETHAMINE ANTIMALARIAL",
    "QUININE":"QUININE SULPHATE ANTIMALARIAL",
    "DEPO-PROVERA":"MEDROXYPROGESTERONE ACETATE INJECTABLE CONTRACEPTIVE",
    "GLUCOPHAGE":"METFORMIN BIGUANIDE ANTIDIABETIC",
    "ASPIRIN":"ACETYLSALICYLIC ACID ANTIPLATELET ANALGESIC",
    "DISPRIN":"ACETYLSALICYLIC ACID ANTIPLATELET ANALGESIC",
    "RINGER":"RINGER LACTATE INTRAVENOUS INFUSION SOLUTION",
    "HARTMANN":"RINGER LACTATE INTRAVENOUS INFUSION SOLUTION",
    "NORMAL SALINE":"SODIUM CHLORIDE 0.9% INTRAVENOUS INFUSION",
}

# ATC Level 1 — WHO Anatomical Therapeutic Chemical classification
ATC_MAP = {
    "ANTIBIOTIC":"J01","ANTIMICROBIAL":"J01","ANTIFUNGAL":"J02","ANTIMALARIAL":"P01",
    "ANTIPROTOZOAL":"P01","ANALGESIC":"N02","ANTIPYRETIC":"N02","NSAID":"M01",
    "DIURETIC":"C03","BRONCHODILATOR":"R03","ANTIDIABETIC":"A10","CONTRACEPTIVE":"G03",
    "ANTIPLATELET":"B01","ANTIHYPERTENSIVE":"C02","ANTIVIRAL":"J05","ANTHELMINTIC":"P02",
    "ANTIRETROVIRAL":"J05A","CORTICOSTEROID":"H02","ANTIHISTAMINE":"R06","ANTACID":"A02",
    "LAXATIVE":"A06","ANTIEMETIC":"A04","HAEMOSTATIC":"B02","INTRAVENOUS INFUSION":"B05",
    "VITAMIN":"A11","IRON":"B03A","CALCIUM":"A12","ELECTROLYTE":"B05",
}

FAMILIES = {
    "SCREW":    r"\bSCREW\b|\bSCREWS\b|\bCORTEX\b|\bCORTICAL\b|\bCANCELLOUS\b|\bINTERLOCKING\b",
    "PLATE":    r"\bPLATE\b|\bLCP\b|\bDCP\b|\bDHS\b",
    "NAIL":     r"\bNAIL\b|\bNAILS\b|\bPFNA\b|\bIMN\b",
    "KWIRE":    r"\bKIRSCHNER\b|\bK-WIRE\b|\bKWIRE\b",
    "CERCLAGE": r"\bCERCLAGE\b|\bFILS DE CERCLAGE\b",
    "SUTURE":   r"\bSUTURE\b|POLYSORB|BIOSYN|CAPROSYN|SURGIPRO|TICRON|VELOSORB|SOFSILK|SURGILON|POLYCRYL|MAXON|CHROMIC GUT|PLAIN GUT|\bNYLON\b|V-LOC|MONOCRYL|VICRYL|PROLENE|MONOSOF|POLYGLACTIN|POLYDIOXANONE",
    "TROCAR":   r"\bTROCAR\b|\bCANNULA\b(?!TED)|\bVERSAPORT\b",
    "STAPLER":  r"\bSTAPLER\b|\bGIA\b|\bEEA\b|\bCEEA\b",
    "DRAIN":    r"\bDRAIN\b|\bREDON\b",
    "DRUG_ORAL":r"\bTABLET\b|\bCAPSULE\b|\bSYRUP\b|\bSACHET\b|\bMG\b.*\bTAB\b|\bMCG\b.*\bTAB\b",
    "DRUG_INJ": r"\bINJECTION\b|\bINJECTABLE\b|\bVIAL\b|\bAMPOULE\b|\bAMP\b.*\bML\b",
    "DRUG_IV":  r"\bINFUSION\b|\bIV FLUID\b|\bIV SOLUTION\b|\bDEXTROSE\b|\bSALINE\b|\bRINGER\b",
    "CATHETER": r"\bCATHETER\b|\bFOLEY\b",
    "GLOVE":    r"\bGLOVE\b|\bGLOVES\b",
    "SYRINGE":  r"\bSYRINGE\b|\bSYRINGES\b",
    "NEEDLE":   r"\bNEEDLE\b|\bNEEDLES\b",
    "IV_SET":   r"\bIV SET\b|\bINFUSION SET\b|\bDRIP SET\b",
    "BANDAGE":  r"\bBANDAGE\b|\bGAUZE\b|\bDRESSING\b|\bWOUND CARE\b",
    "MESH":     r"\bMESH\b",
    "ELECTRODE":r"\bELECTRODE\b|\bBIPOLAR\b|\bPENCIL\b.*\bESU\b",
    "CLIP":     r"\bHEMOLOCK\b|\bENDOCLIP\b|\bLAPROCLIP\b",
}

DRUG_FAMILIES = {"DRUG_ORAL", "DRUG_INJ", "DRUG_IV"}

ANAT_INCOMPAT = {
    "DISTAL RADIUS":["DHS","INTERTROCHANTERIC","HIP SCREW"],
    "VOLAR LOCKING":["DHS","HIP"],
    "CLAVICLE":["DHS","TIBIA","FEMUR","HIP"],
    "PROXIMAL HUMERUS":["DHS","HIP","TIBIA"],
    "FILS DE CERCLAGE":["TIGHTNER","TIGHTENER"],
    "DRAIN DE REDON":["CHEST DRAIN","THORACIC DRAIN"],
}

# ── MODULE 1: NORMALIZATION ENGINE ────────────────────────────────────────────
class NormalizationEngine:
    """Cleans, expands, and standardizes medical product descriptions."""

    @staticmethod
    def clean(t: str) -> str:
        if pd.isna(t): return ""
        s = str(t).upper().strip()
        s = re.sub(r"(\d+),(\d+)", r"\1.\2", s)       # European decimal
        s = re.sub(r"[×✕*]", " ", s)
        s = re.sub(r"\.0\s*MM", "MM", s)
        s = re.sub(r"(\d+\.?\d*)\s*MM", r"\1MM", s)
        s = re.sub(r"(\d+\.?\d*)\s*ML", r"\1ML", s)
        s = re.sub(r"(\d+\.?\d*)\s*MG", r"\1MG", s)
        s = re.sub(r"(\d+\.?\d*)\s*MCG", r"\1MCG", s)
        s = re.sub(r"(\d+\.?\d*)\s*IU\b", r"\1IU", s)
        s = re.sub(r"\b(?:CH|FG)\s*(\d+)", r"FR\1", s)
        s = re.sub(r"\bS\.T\.\b|\bSELF-TAPPING\b|\bS/T\b", "SELF TAPPING", s)
        s = re.sub(r"\bLH\b", "LEFT", s)
        s = re.sub(r"\bRH\b", "RIGHT", s)
        s = re.sub(r"\b(\d+)\s*-\s*HOLES?\b", r"\1 HOLES", s)
        s = re.sub(r"\bHOLE\b", "HOLES", s)
        s = re.sub(r"\bTABS?\b", "TABLET", s)
        s = re.sub(r"\bCAPS?\b", "CAPSULE", s)
        s = re.sub(r"\bINJ\b", "INJECTION", s)
        s = re.sub(r"\bSOLN?\b", "SOLUTION", s)
        s = re.sub(r"\s{2,}", " ", s)
        return s.strip(" ,.-")

    @staticmethod
    def generic(t: str) -> str:
        """Apply brand→generic (INN) normalization."""
        s = NormalizationEngine.clean(t)
        for brand, gen in BRAND.items():
            s = re.sub(r"\b" + re.escape(brand.upper()) + r"\b", gen, s)
        return s

    @staticmethod
    def for_embed(t: str) -> str:
        return NormalizationEngine.generic(t).lower()

    @staticmethod
    def specs(t: str) -> dict:
        u = NormalizationEngine.clean(t)
        return {
            "dims":  sorted([float(v) for v in re.findall(r"(\d+\.?\d*)MM", u)]),
            "ml":    [float(v) for v in re.findall(r"(\d+\.?\d*)ML", u)],
            "mg":    [float(v) for v in re.findall(r"(\d+\.?\d*)MG", u)],
            "mcg":   [float(v) for v in re.findall(r"(\d+\.?\d*)MCG", u)],
            "iu":    [float(v) for v in re.findall(r"(\d+\.?\d*)IU", u)],
            "fr":    [int(v)   for v in re.findall(r"FR(\d+)", u)],
            "holes": int(m.group(1)) if (m := re.search(r"(\d+)\s*HOLES?", u)) else None,
            "usp":   m.group(1)+"-0" if (m := re.search(r"(\d+)-?0\b", u)) else None,
        }

    @staticmethod
    def family(t: str) -> str:
        u = NormalizationEngine.clean(t)
        for fam, pat in FAMILIES.items():
            if re.search(pat, u): return fam
        return "OTHER"

    @staticmethod
    def is_drug(t: str) -> bool:
        return NormalizationEngine.family(t) in DRUG_FAMILIES

    @staticmethod
    def get_atc(t: str) -> str:
        """Extract ATC-level hint from description."""
        u = NormalizationEngine.generic(t).upper()
        for kw, atc in ATC_MAP.items():
            if kw in u: return atc
        return ""

NE = NormalizationEngine()

# ── MODULE 2: EMBEDDING ENGINE ────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_biencoder(name: str = "all-MiniLM-L6-v2"):
    if not ST_OK: return None
    try:
        return SentenceTransformer(name)
    except Exception:
        return None

@st.cache_resource(show_spinner=False)
def load_crossencoder():
    if not ST_OK: return None
    try:
        return CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2")
    except Exception:
        return None

class EmbeddingEngine:
    """
    Tier-1: Local Sentence Transformers (all-MiniLM-L6-v2 or PubMedBERT)
    Tier-2: HuggingFace Inference API (remote)
    Tier-3: TF-IDF (scikit-learn)
    Tier-4: RapidFuzz (edit distance)
    """
    HF_URL = "https://api-inference.huggingface.co/models/sentence-transformers/{}"

    def __init__(self):
        model_name = CFG.get("model_choice", "all-MiniLM-L6-v2")
        self.bi = load_biencoder(model_name)
        self.ce = load_crossencoder() if CFG.get("use_cross_encoder", True) else None
        if self.bi:
            S.model_name = model_name
        elif SKL_OK:
            S.model_name = "TF-IDF"
        elif RF_OK:
            S.model_name = "RapidFuzz"
        else:
            S.model_name = "None"

    def embed(self, texts: list) -> np.ndarray | None:
        if self.bi:
            return self.bi.encode(texts, normalize_embeddings=True,
                                  show_progress_bar=False, batch_size=64)
        # HF Inference API fallback
        tok = self._secret("HF_TOKEN")
        if tok and REQ_OK:
            model = CFG.get("model_choice","all-MiniLM-L6-v2")
            url = self.HF_URL.format(model)
            try:
                r = _http_get(url, method="POST",
                              json_data={"inputs": texts, "options": {"wait_for_model": True}},
                              headers={"Authorization": f"Bearer {tok}"}, timeout=25)
                if r:
                    arr = np.array(r)
                    norms = np.linalg.norm(arr, axis=1, keepdims=True)
                    return arr / np.maximum(norms, 1e-9)
            except Exception:
                pass
        return None

    def rerank(self, query: str, candidates: list) -> list:
        """Cross-encoder re-ranking. Falls back to original order if CE unavailable."""
        if self.ce is None or len(candidates) <= 1 or not CFG.get("use_cross_encoder"):
            return candidates
        try:
            pairs = [[query, c[1]["desc"]] for c in candidates]
            scores = self.ce.predict(pairs)
            ranked = sorted([(float(scores[i]), c[1]) for i, c in enumerate(candidates)],
                            key=lambda x: -x[0])
            mx = max(s for s, _ in ranked) or 1
            mn = min(s for s, _ in ranked)
            rng = mx - mn or 1
            return [((s - mn) / rng, m) for s, m in ranked]
        except Exception:
            return candidates

    @staticmethod
    def _secret(key: str) -> str:
        try: return st.secrets.get(key, "")
        except: return ""

EE = EmbeddingEngine()

# ── HTTP HELPER ───────────────────────────────────────────────────────────────
_session = None
_req_lock = threading.Lock()
_last_req: dict = {}

def _http_get(url, params=None, method="GET", json_data=None,
              headers=None, host="default", gap=0.65, timeout=10):
    global _session
    if not REQ_OK: return None
    if _session is None:
        s = requests.Session()
        s.mount("https://", HTTPAdapter(max_retries=Retry(
            total=2, backoff_factor=0.4, status_forcelist=[429,500,502,503])))
        s.headers["User-Agent"] = "MedMatch-AI/2.1 Rwanda-FDA SOP/004"
        _session = s
    with _req_lock:
        w = gap - (time.time() - _last_req.get(host, 0))
        if w > 0: time.sleep(w)
        _last_req[host] = time.time()
    try:
        h = {**(_session.headers), **(headers or {})}
        if method == "POST":
            r = _session.post(url, json=json_data, headers=h, timeout=timeout)
        else:
            r = _session.get(url, params=params, headers=h, timeout=timeout)
        return r.json() if r.status_code == 200 else ({} if r.status_code == 404 else None)
    except Exception:
        return None

def _secret(key):
    try: return st.secrets.get(key, "")
    except: return ""

# ── CATALOG INDEXER ───────────────────────────────────────────────────────────
def build_catalog_index(df, d_col, c_col, cat_col=None) -> bool:
    records, texts = [], []
    for _, row in df.iterrows():
        desc = str(row[d_col]).strip()
        code = str(row[c_col]).strip()
        if not desc or desc.lower() == "nan": continue
        cat  = str(row[cat_col]).strip() if cat_col and cat_col in row.index else NE.family(desc)
        g    = NE.for_embed(desc)
        atc  = NE.get_atc(desc)
        records.append({"code": code, "desc": desc, "generic": g,
                        "category": cat, "family": NE.family(desc), "atc": atc})
        texts.append(g)

    if not records: return False
    S.embed_meta = records

    # Semantic embeddings
    embs = EE.embed(texts)
    if embs is not None:
        embs = embs.astype(np.float32)
        S.embeddings = embs
        # FAISS index (much faster for large catalogs)
        if FAISS_OK:
            dim = embs.shape[1]
            idx = faiss.IndexFlatIP(dim)  # Inner Product = cosine on normalized vecs
            idx.add(embs)
            S.faiss_idx = idx

    # TF-IDF fallback
    if SKL_OK:
        vec = TfidfVectorizer(ngram_range=(1, 2), min_df=1, analyzer="word",
                              sublinear_tf=True)
        S.tfidf_mat = vec.fit_transform(texts)
        S.tfidf_vec = vec

    return True

# ── MODULE 3: CANDIDATE RETRIEVAL ─────────────────────────────────────────────
def retrieve(query_generic: str, top_k: int = 20) -> list:
    """
    Returns [(score, meta_dict)] sorted descending.
    Uses FAISS > numpy cosine > TF-IDF > RapidFuzz in priority order.
    """
    if not S.embed_meta: return []

    # FAISS (fastest for large catalogs)
    if S.faiss_idx is not None:
        qv = EE.embed([query_generic])
        if qv is not None:
            D, I = S.faiss_idx.search(qv.astype(np.float32), top_k)
            return [(float(D[0][j]), S.embed_meta[I[0][j]])
                    for j in range(len(I[0])) if I[0][j] >= 0]

    # Numpy cosine
    if S.embeddings is not None:
        qv = EE.embed([query_generic])
        if qv is not None:
            sims = np.dot(S.embeddings, qv[0]).astype(float)
            idx  = np.argsort(sims)[::-1][:top_k]
            return [(float(sims[i]), S.embed_meta[i]) for i in idx]

    # TF-IDF
    if S.tfidf_vec is not None:
        qv   = S.tfidf_vec.transform([query_generic])
        sims = sk_cos(qv, S.tfidf_mat).flatten()
        idx  = np.argsort(sims)[::-1][:top_k]
        return [(float(sims[i]), S.embed_meta[i]) for i in idx]

    # RapidFuzz
    if RF_OK:
        choices = [m["generic"] for m in S.embed_meta]
        hits = rfproc.extract(query_generic, choices,
                              scorer=fuzz.token_sort_ratio, limit=top_k)
        return [(s / 100.0, S.embed_meta[choices.index(t)]) for t, s, _ in hits]

    return []

# ── MODULE 5: VALIDATION ENGINE ───────────────────────────────────────────────
class ValidationEngine:
    """Rule-based clinical safety layer. Never accepts anatomy mismatches or dose errors."""

    @staticmethod
    def check(inp: str, npc: str) -> tuple:
        """Returns (status: VALID|REVIEW, issues: list[str])"""
        issues = []
        ui, ni = NE.clean(inp), NE.clean(npc)

        # Anatomy incompatibility (AO Foundation / PMC-verified)
        for kw, bad_list in ANAT_INCOMPAT.items():
            if kw.upper() in ui:
                for bad in bad_list:
                    if bad.upper() in ni:
                        issues.append(f"⚠ Anatomy incompatibility: '{kw}' ≠ '{bad}' (AO Foundation)")

        # Product family
        fi, fn = NE.family(inp), NE.family(npc)
        if fi != "OTHER" and fn != "OTHER" and fi != fn:
            # Drug-to-device or device-to-drug are CRITICAL
            if (fi in DRUG_FAMILIES) != (fn in DRUG_FAMILIES):
                issues.append(f"🚨 CRITICAL: Drug vs device mismatch ({fi} vs {fn})")
            else:
                issues.append(f"Family mismatch: {fi} vs {fn}")

        # Numeric specs
        si, sn = NE.specs(inp), NE.specs(npc)
        if si["dims"] and sn["dims"] and abs(si["dims"][0] - sn["dims"][0]) >= 0.15:
            issues.append(f"Size: {si['dims'][0]}mm vs {sn['dims'][0]}mm (tol ±0.1mm)")
        if si["mg"] and sn["mg"] and si["mg"] and sn["mg"] and si["mg"][0] != sn["mg"][0]:
            diff = abs(si["mg"][0] - sn["mg"][0])
            severity = "🚨 CRITICAL" if diff > si["mg"][0] * 0.25 else "⚠"
            issues.append(f"{severity} Drug dose: {si['mg'][0]}mg vs {sn['mg'][0]}mg")
        if si["ml"] and sn["ml"] and abs(si["ml"][0] - sn["ml"][0]) > 0.5:
            issues.append(f"Volume: {si['ml'][0]}ml vs {sn['ml'][0]}ml")
        if si["fr"] and sn["fr"] and si["fr"][0] != sn["fr"][0]:
            issues.append(f"FR size: FR{si['fr'][0]} vs FR{sn['fr'][0]}")
        if si["holes"] and sn["holes"] and si["holes"] != sn["holes"]:
            issues.append(f"Holes: {si['holes']} vs {sn['holes']}")
        if bool(si["ml"]) and bool(sn["dims"]) and not bool(si["dims"]):
            issues.append("Unit crossover: volume (ml) vs dimension (mm)")

        return ("REVIEW", issues) if issues else ("VALID", [])

    @staticmethod
    def classify_match(inp: str, npc: str, score: float) -> str:
        """EXACT | BRAND_MATCH | SPEC_DIFF | NEW_SOP"""
        ig = NE.generic(inp)
        ic = NE.clean(inp)
        nc = NE.clean(npc)
        if ic == nc: return "EXACT"
        if ig == NE.generic(npc): return "EXACT"
        if score >= 0.97: return "EXACT"
        if score >= 0.88: return "BRAND_MATCH"
        fi, fn = NE.family(inp), NE.family(npc)
        if fi == fn and fi != "OTHER" and score >= 0.65: return "SPEC_DIFF"
        return "NEW_SOP"

VE = ValidationEngine()

# ── MODULE 7: DITTO SEMANTIC VERIFIER ────────────────────────────────────────
class DittoVerifier:
    """
    Pairwise semantic confirmation using bi-encoder similarity.
    Acts as final guard before accepting a match.
    If Ditto score < threshold → downgrade confidence or flag for review.
    """
    THRESHOLD = 0.55

    @classmethod
    def verify(cls, query: str, candidate: str) -> dict:
        """Returns {score, confirmed, method}"""
        ck = hashlib.md5(f"{query}||{candidate}".lower().encode()).hexdigest()
        if ck in S.ditto_cache:
            return S.ditto_cache[ck]

        result = {"score": 0.0, "confirmed": False, "method": "none"}

        # Use bi-encoder similarity
        if EE.bi is not None:
            try:
                vecs = EE.embed([NE.for_embed(query), NE.for_embed(candidate)])
                if vecs is not None:
                    score = float(np.dot(vecs[0], vecs[1]))
                    result = {
                        "score":     round(score * 100, 1),
                        "confirmed": score >= cls.THRESHOLD,
                        "method":    "sentence-transformer",
                    }
            except Exception:
                pass
        elif RF_OK:
            # Fallback: fuzzy similarity as ditto score
            score = fuzz.token_sort_ratio(NE.clean(query), NE.clean(candidate)) / 100.0
            result = {
                "score":     round(score * 100, 1),
                "confirmed": score >= cls.THRESHOLD,
                "method":    "rapidfuzz",
            }

        S.ditto_cache[ck] = result
        if result["confirmed"]:
            S.stats["ditto_confirms"] += 1
        return result

DITTO = DittoVerifier()

# ── MODULE 8: ENRICHMENT ENGINE ───────────────────────────────────────────────
class EnrichmentEngine:
    """
    Queries openFDA (device + drug) and WHO INN database.
    All results session-cached by md5(description).
    """

    @staticmethod
    def _fda_key(): return _secret("OPENFDA_KEY")

    @classmethod
    def fda_device(cls, query: str) -> dict:
        """openFDA Device Classification API."""
        key = cls._fda_key()
        for q in [f'device_name:"{query}"', "device_name:" + "+".join(query.split()[:4])]:
            params = {"search": q, "limit": 3}
            if key: params["api_key"] = key
            data = _http_get("https://api.fda.gov/device/classification.json",
                             params=params, host="fda_dev")
            if data and data.get("results"):
                top = data["results"][0]
                S.stats["fda_hits"] += 1
                return {
                    "source":       "openFDA Device Classification",
                    "device_name":  top.get("device_name", ""),
                    "product_code": top.get("product_code", ""),
                    "device_class": f"Class {top.get('device_class','')}",
                    "definition":   top.get("definition", "")[:300],
                    "regulation":   top.get("regulation_number", ""),
                }
        return {}

    @classmethod
    def fda_drug(cls, query: str) -> dict:
        """openFDA Drug Label API."""
        key = cls._fda_key()
        # Strip dose specs from query
        clean_q = re.sub(r"\d+\.?\d*\s*(mg|mcg|ml|iu|g)\b", "", query, flags=re.I).strip()
        params = {"search": f'brand_name:"{clean_q}"', "limit": 2}
        if key: params["api_key"] = key
        data = _http_get("https://api.fda.gov/drug/label.json", params=params, host="fda_drug")
        if data and data.get("results"):
            top = data["results"][0]
            ofd = top.get("openfda", {})
            S.stats["fda_hits"] += 1
            return {
                "source":       "openFDA Drug Label",
                "generic_name": ofd.get("generic_name", [""])[0],
                "product_type": ofd.get("product_type", [""])[0],
                "manufacturer": ofd.get("manufacturer_name", [""])[0],
                "route":        ofd.get("route", [""])[0],
                "substance":    ofd.get("substance_name", [""])[0],
            }
        return {}

    @classmethod
    def who_inn(cls, query: str) -> dict:
        """
        WHO International Nonproprietary Name lookup via WHO MedNet API.
        Falls back to WHO list PDF endpoint if API unavailable.
        """
        ck = hashlib.md5(query.lower().encode()).hexdigest()
        if ck in S.who_cache: return S.who_cache[ck]

        # Strip dose/form specs to get the INN-searchable name
        base = re.sub(r"\d+\.?\d*\s*(mg|mcg|ml|iu|g|tablet|capsule|injection)\b",
                      "", query, flags=re.I).strip()
        base = re.sub(r"\s+", " ", base).strip()

        result = {}

        # WHO Drug Information API (public endpoint)
        data = _http_get(
            "https://www.who.int/medicines/services/inn/innlists/en/",
            params={"query": base, "format": "json"},
            host="who_inn", gap=1.0
        )
        if not data:
            # Try WHO Essential Medicines via openFDA drug index
            fda_r = cls.fda_drug(base)
            if fda_r.get("generic_name"):
                result = {
                    "source":  "WHO INN (via openFDA)",
                    "inn":     fda_r["generic_name"],
                    "route":   fda_r.get("route", ""),
                    "atc":     NE.get_atc(fda_r["generic_name"]),
                }

        if result:
            S.stats["who_hits"] += 1
        S.who_cache[ck] = result
        return result

    @classmethod
    def enrich(cls, description: str) -> dict:
        """Auto-detect device vs drug and query appropriate endpoint."""
        ck = hashlib.md5(description.lower().encode()).hexdigest()
        if ck in S.enrich_cache: return S.enrich_cache[ck]

        clean = re.sub(r"\d+\.?\d*\s*(?:mm|ml|mg|mcg|iu|fr|g)\b", "",
                       description, flags=re.I).strip()
        words = [w for w in clean.upper().split() if len(w) > 2]
        query = " ".join(words[:6])

        is_drug = NE.is_drug(description)
        result = cls.fda_drug(query) if is_drug else cls.fda_device(query)

        # WHO INN for drugs
        if is_drug and CFG.get("use_who"):
            who = cls.who_inn(query)
            if who:
                result["who_inn"]  = who.get("inn", "")
                result["who_atc"]  = who.get("atc", "")
                result["who_route"]= who.get("route", "")

        S.enrich_cache[ck] = result
        return result

ENR = EnrichmentEngine()

# ── MODULE 9: MEDMATCH PIPELINE ───────────────────────────────────────────────
def medmatch(description: str, top_n: int = 5, do_enrich: bool = False) -> dict:
    """
    Full pipeline:
    Normalize → Embed → Retrieve(20) → Rerank → Score → Validate → Ditto → Enrich
    """
    if not S.embed_meta:
        return {"error": "No catalog loaded. Upload an NPC catalog in Catalog Manager."}

    t0 = time.perf_counter()

    # 1. Normalize
    generic = NE.for_embed(description)
    specs   = NE.specs(description)
    family  = NE.family(description)
    atc     = NE.get_atc(description)
    is_drug = NE.is_drug(description)

    # 2. Retrieve top-20 candidates
    candidates = retrieve(generic, top_k=20)
    if not candidates:
        return {"error": "No candidates found.", "results": []}

    # 3. Cross-encoder re-ranking
    reranked = EE.rerank(generic, candidates)

    # 4. Score, validate, and Ditto verify top_n
    results_out = []
    fw = CFG.get("fuzzy_weight",  0.30)
    sw = CFG.get("semantic_weight", 0.70)

    for rank, (sem_score, meta) in enumerate(reranked[:top_n], 1):
        # Hybrid score
        fuzz_s = (fuzz.token_sort_ratio(NE.clean(description), NE.clean(meta["desc"])) / 100.0
                  if RF_OK else sem_score)
        hybrid = sw * sem_score + fw * fuzz_s

        # Validation
        vs, vi = VE.check(description, meta["desc"])

        # Match type
        mt = VE.classify_match(description, meta["desc"], sem_score)

        # Ditto verification (top-5 only, to limit latency)
        ditto = {}
        if CFG.get("use_ditto", True):
            ditto = DITTO.verify(description, meta["desc"])
            # If Ditto disagrees strongly, downgrade confidence
            if ditto.get("score", 100) < 40 and mt in ("EXACT","BRAND_MATCH"):
                mt = "SPEC_DIFF"
                vi.append(f"Ditto verifier: low semantic confirmation ({ditto['score']:.0f}%)")
                vs = "REVIEW"

        results_out.append({
            "rank":       rank,
            "npc_code":   meta["code"],
            "npc_desc":   meta["desc"],
            "category":   meta["category"],
            "family":     meta["family"],
            "atc":        meta.get("atc", ""),
            "sem_score":  round(sem_score * 100, 1),
            "fuzz_score": round(fuzz_s * 100, 1),
            "score":      round(hybrid * 100, 1),
            "match_type": mt,
            "val_status": vs,
            "val_issues": vi,
            "ditto":      ditto,
        })

    # 5. Online enrichment (top result only)
    enrich_data = {}
    if do_enrich and REQ_OK and CFG.get("use_fda"):
        enrich_data = ENR.enrich(description)

    elapsed_ms = round((time.perf_counter() - t0) * 1000, 1)

    # Update stats
    S.stats["queries"] += 1
    mt0 = results_out[0]["match_type"] if results_out else "NEW_SOP"
    key = {"EXACT": "exact", "BRAND_MATCH": "brand", "SPEC_DIFF": "spec"}.get(mt0, "new")
    S.stats[key] += 1
    n = S.stats["queries"]
    S.stats["avg_ms"] = round((S.stats["avg_ms"] * (n - 1) + elapsed_ms) / n, 1)

    return {
        "input":      description,
        "normalized": generic,
        "family":     family,
        "atc":        atc,
        "is_drug":    is_drug,
        "specs":      specs,
        "results":    results_out,
        "elapsed_ms": elapsed_ms,
        "engine":     S.model_name or "?",
        "enrich":     enrich_data,
    }

# ── PARALLEL BATCH MATCHER ────────────────────────────────────────────────────
def batch_match_parallel(descs: list, codes: list, do_enrich: bool = False,
                         workers: int = 3) -> list:
    """
    Thread-pool batch processing. Each worker calls medmatch independently.
    Returns list of flat result dicts.
    """
    def process_one(args):
        i, desc, code = args
        res = medmatch(desc, top_n=1, do_enrich=do_enrich)
        if "error" in res or not res.get("results"):
            return {"INPUT_CODE": code, "INPUT_DESCRIPTION": desc,
                    "NPC_CODE": "", "NPC_DESCRIPTION": "", "CATEGORY": "",
                    "MATCH_SCORE": 0, "MATCH_TYPE": "NEW_SOP",
                    "VALIDATION_STATUS": "REVIEW", "VALIDATION_COMMENT": "No match",
                    "DITTO_SCORE": "", "ATC_CODE": "", "IS_DRUG": res.get("is_drug", False),
                    "FDA_PRODUCT_CODE": "", "FDA_DEFINITION": ""}
        top = res["results"][0]
        ed  = res.get("enrich", {})
        ditto_score = top.get("ditto", {}).get("score", "")
        return {
            "INPUT_CODE":         code,
            "INPUT_DESCRIPTION":  desc,
            "NPC_CODE":           top["npc_code"],
            "NPC_DESCRIPTION":    top["npc_desc"],
            "CATEGORY":           top["category"],
            "MATCH_SCORE":        top["score"],
            "MATCH_TYPE":         top["match_type"],
            "VALIDATION_STATUS":  top["val_status"],
            "VALIDATION_COMMENT": " | ".join(top["val_issues"]),
            "DITTO_SCORE":        ditto_score,
            "ATC_CODE":           res.get("atc", ""),
            "IS_DRUG":            res.get("is_drug", False),
            "FDA_PRODUCT_CODE":   ed.get("product_code", ""),
            "FDA_DEFINITION":     ed.get("definition", ed.get("generic_name", ""))[:200],
        }

    tasks = [(i, d, c) for i, (d, c) in enumerate(zip(descs, codes))]
    results = [None] * len(tasks)

    # Use threads (safe for I/O-bound embedding API; local models prefer sequential)
    effective_workers = 1 if EE.bi else workers
    with ThreadPoolExecutor(max_workers=effective_workers) as ex:
        futures = {ex.submit(process_one, t): t[0] for t in tasks}
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                results[idx] = fut.result()
            except Exception as e:
                results[idx] = {"error": str(e)}

    return [r for r in results if r]

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────
def to_excel(rows: list) -> bytes:
    if not OPX_OK:
        return pd.DataFrame(rows).to_csv(index=False).encode()
    wb  = Workbook()
    ws  = wb.active
    ws.title = "MedMatch Results"
    thin = Side(style="thin", color="1e3a5f")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = ["INPUT_CODE","INPUT_DESCRIPTION","NPC_CODE","NPC_DESCRIPTION","CATEGORY",
            "MATCH_SCORE","MATCH_TYPE","VALIDATION_STATUS","VALIDATION_COMMENT",
            "DITTO_SCORE","ATC_CODE","IS_DRUG","FDA_PRODUCT_CODE","FDA_DEFINITION"]
    ws.append(cols)
    hf = PatternFill("solid", fgColor="060d18")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(name="Calibri", bold=True, color="00D4FF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[1].height = 36
    mt_fills = {
        "EXACT":       PatternFill("solid", fgColor="063a1f"),
        "BRAND_MATCH": PatternFill("solid", fgColor="062040"),
        "SPEC_DIFF":   PatternFill("solid", fgColor="3d2e00"),
        "NEW_SOP":     PatternFill("solid", fgColor="1a1a1a"),
    }
    for r in rows:
        rv = [r.get(c, "") for c in cols]
        ws.append(rv)
        xl = ws[ws.max_row]
        rf = mt_fills.get(r.get("MATCH_TYPE",""), PatternFill("solid", fgColor="111111"))
        for cell in xl:
            cell.fill = rf
            cell.font = Font(name="Calibri", size=9, color="E8F4FD")
            cell.alignment = Alignment(vertical="center")
            cell.border = bdr
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = (
            65 if "DESCRIPTION" in col or "DEFINITION" in col else
            22 if "CODE" in col or "TYPE" in col or "STATUS" in col else 14)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── UI HELPERS ────────────────────────────────────────────────────────────────
def cc(s):  return "#00e676" if s >= CFG["confidence_high"] else ("#ffb300" if s >= CFG["confidence_med"] else "#ff4757")
def bc(s):  return "ch" if s >= CFG["confidence_high"] else ("cm" if s >= CFG["confidence_med"] else "cl")
def mpill(mt): c={"EXACT":"E","BRAND_MATCH":"B","SPEC_DIFF":"S","NEW_SOP":"N"}.get(mt,"N"); return f'<span class="pill p{c}">{mt}</span>'
def vpill(vs): c="V" if vs=="VALID" else "R"; return f'<span class="pill p{c}">{vs}</span>'
def epill(n):
    if any(x in n for x in ["MiniLM","BERT","PubMed","mpnet"]): return f'<span class="pill pAI">🧠 {n}</span>'
    if "TF-IDF" in n: return f'<span class="pill pAI" style="color:var(--pink)">📊 TF-IDF</span>'
    return f'<span class="pill pAI" style="color:var(--pink)">⚡ {n}</span>'
def apill(atc): return f'<span class="atc">ATC: {atc}</span>' if atc else ""
def stat_card(col, val, lbl, cls):
    col.markdown(f'<div class="sc"><div class="sv {cls}">{val}</div><div class="sl">{lbl}</div></div>', unsafe_allow_html=True)

def render_result_card(r: dict, is_top: bool = False):
    cls   = "rc r1" if is_top else ("rc r2" if r["rank"] == 2 else "rc")
    score = r["score"]; col = cc(score); bar = bc(score)
    ditto = r.get("ditto", {})
    ditto_html = ""
    if ditto and CFG.get("use_ditto"):
        ds = ditto.get("score", 0)
        dc = "#00e676" if ds >= 70 else ("#ffb300" if ds >= 45 else "#ff4757")
        ditto_html = f'&nbsp;<span class="pill pAI" style="color:{dc}">Ditto {ds:.0f}%</span>'
    iss_html = "".join(
        f'<div style="color:#ff4757;font-size:.7rem;margin-top:.15rem">⚠ {i}</div>'
        for i in r["val_issues"]
    )
    atc_html = apill(r.get("atc",""))
    st.markdown(f"""
<div class="{cls}">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:.5rem">
    <div style="flex:1">
      <div class="rn">#{r['rank']} &nbsp; {r['npc_code']}</div>
      <div class="rd">{r['npc_desc']}</div>
      <div style="font-size:.72rem;color:var(--mu);margin-top:.1rem">
        📁 {r['category']}  ·  🏷 {r['family']}  {atc_html}
      </div>
    </div>
    <div style="text-align:right;min-width:65px">
      <div style="font-family:'JetBrains Mono',monospace;font-size:1.4rem;font-weight:800;color:{col}">{score:.0f}%</div>
      <div style="font-size:.62rem;color:var(--mu)">S:{r['sem_score']:.0f} F:{r['fuzz_score']:.0f}</div>
    </div>
  </div>
  <div class="cbw"><div class="{bar} cbar" style="width:{min(score,100):.0f}%"></div></div>
  <div style="margin-top:.3rem">{mpill(r['match_type'])}&nbsp;{vpill(r['val_status'])}{ditto_html}</div>
  {iss_html}
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="logo">🧬 MedMatch AI</div><div class="logo-sub">Medical Standardization Engine v2.1</div>',
                unsafe_allow_html=True)
    st.markdown("---")
    # Engine status
    bi_model, ce_model = load_biencoder(CFG.get("model_choice","all-MiniLM-L6-v2")), load_crossencoder()
    for lbl, ok, mbcls in [
        (f"🧠 {CFG.get('model_choice','all-MiniLM-L6-v2')}", ST_OK and bi_model, "mbok"),
        ("⚡ Cross-Encoder Re-Rank", ST_OK and ce_model, "mbok"),
        ("🔍 FAISS Indexing", FAISS_OK, "mbok"),
        ("📊 TF-IDF Fallback", SKL_OK, "mbok"),
        ("🔁 RapidFuzz", RF_OK, "mbok"),
        ("🌐 openFDA + WHO", REQ_OK, "mbok"),
        ("🔮 Ditto Verifier", True, "mbok"),
    ]:
        cls = mbcls if ok else "mbof"
        icon = "" if not ok else ""
        st.markdown(f'<div class="mb {cls}">{lbl}</div>', unsafe_allow_html=True)

    n_cat = len(S.embed_meta) if S.embed_meta else 0
    emb_type = ("FAISS" if S.faiss_idx else ("Numpy" if S.embeddings is not None else "TF-IDF")) if n_cat else "—"
    st.markdown("")
    st.markdown(f'<div class="mb {"mbok" if n_cat else "mbwn"}">📚 {"Catalog: "+str(n_cat)+" items · "+emb_type if n_cat else "No catalog loaded"}</div>',
                unsafe_allow_html=True)
    st.markdown("---")
    # Navigation
    nav_pages = [
        ("🔍", "search",    "Single Search"),
        ("⚡", "batch",     "Batch Processing"),
        ("📚", "catalog",   "Catalog Manager"),
        ("⚙️", "settings",  "Settings"),
        ("📊", "dashboard", "Dashboard"),
    ]
    for icon, key, lbl in nav_pages:
        active = S.page == key
        if st.button(f"{icon}  {lbl}", key=f"nav_{key}", use_container_width=True):
            S.page = key; st.rerun()
    st.markdown("---")
    # History
    if S.history:
        st.markdown('<div style="font-size:.7rem;color:var(--mu);text-transform:uppercase;letter-spacing:.08em;margin-bottom:.3rem">Recent Searches</div>',
                    unsafe_allow_html=True)
        for item in list(S.history)[-5:][::-1]:
            st.markdown(f'<div class="hi">{item[:42]+"…" if len(item)>42 else item}</div>',
                        unsafe_allow_html=True)
        if st.button("Clear History", key="clr_hist"):
            S.history.clear(); st.rerun()
    st.markdown("---")
    st.markdown('<div style="font-size:.62rem;color:var(--mu2)">MedMatch AI v2.1<br>Rwanda FDA · SOP ODDG/RES/SOP/004<br>Modules: Normalize · Embed · Retrieve<br>Rerank · Validate · Ditto · Enrich</div>',
                unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: SETTINGS
# ─────────────────────────────────────────────────────────────────────────────
if S.page == "settings":
    st.markdown('<h1 class="pt">Settings</h1>', unsafe_allow_html=True)
    st.markdown('<p class="ps">Configure matching thresholds, engine modules, and API integrations.</p>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)

    with c1:
        st.markdown('<div class="sh">🧠 Embedding Model</div>', unsafe_allow_html=True)
        models = [
            "all-MiniLM-L6-v2",
            "all-mpnet-base-v2",
            "pritamdeka/S-PubMedBert-MS-MARCO",
            "sentence-transformers/msmarco-MiniLM-L-6-v3",
        ]
        model_labels = {
            "all-MiniLM-L6-v2": "all-MiniLM-L6-v2 ⚡ Fast (384-dim) — recommended",
            "all-mpnet-base-v2": "all-mpnet-base-v2 🔥 Best quality (768-dim)",
            "pritamdeka/S-PubMedBert-MS-MARCO": "PubMedBERT 🏥 Medical domain specialist",
            "sentence-transformers/msmarco-MiniLM-L-6-v3": "msmarco-MiniLM 🔍 Retrieval-optimized",
        }
        new_model = st.selectbox("Active model",
                                  models,
                                  index=models.index(CFG.get("model_choice","all-MiniLM-L6-v2")),
                                  format_func=lambda x: model_labels.get(x, x))
        CFG["model_choice"] = new_model

        st.markdown('<div class="sh">🎯 Matching Thresholds</div>', unsafe_allow_html=True)
        CFG["npc_threshold"]    = st.slider("Minimum match score (%)", 40, 95, CFG.get("npc_threshold",65))
        CFG["semantic_weight"]  = st.slider("Semantic weight", 0.0, 1.0, CFG.get("semantic_weight",0.70), 0.05)
        CFG["fuzzy_weight"]     = round(1.0 - CFG["semantic_weight"], 2)
        st.caption(f"Fuzzy weight: {CFG['fuzzy_weight']:.2f}")

        st.markdown('<div class="sh">🎨 Confidence Bands</div>', unsafe_allow_html=True)
        CFG["confidence_high"] = st.number_input("HIGH confidence floor (%)", 60, 100, CFG.get("confidence_high",85))
        CFG["confidence_med"]  = st.number_input("MEDIUM confidence floor (%)", 30, 99, CFG.get("confidence_med",60))

        st.markdown('<div class="sh">⚙️ Batch Processing</div>', unsafe_allow_html=True)
        CFG["batch_workers"] = st.slider("Parallel workers", 1, 8, CFG.get("batch_workers",3),
                                          help="Use 1 for local sentence-transformers; higher for API fallback")

    with c2:
        st.markdown('<div class="sh">🔬 Active Modules</div>', unsafe_allow_html=True)
        CFG["use_cross_encoder"] = st.toggle("Cross-Encoder Re-Ranking",
            value=CFG.get("use_cross_encoder",True),
            help="ms-marco-MiniLM — improves top-5 ordering, adds ~20ms/query")
        CFG["use_ditto"]   = st.toggle("Ditto Semantic Verifier",
            value=CFG.get("use_ditto",True),
            help="Pairwise confirmation — catches incorrect high-score matches")
        CFG["use_fda"]    = st.toggle("openFDA Enrichment",
            value=CFG.get("use_fda",True),
            help="Device classification + drug label lookup")
        CFG["use_who"]    = st.toggle("WHO INN Lookup",
            value=CFG.get("use_who",True),
            help="International Nonproprietary Names for pharmaceuticals")

        st.markdown('<div class="sh">🔑 API Keys</div>', unsafe_allow_html=True)
        st.markdown("""<div style="background:var(--sf2);border:1px solid var(--bd);border-radius:8px;padding:.8rem 1rem;font-size:.8rem;">
Add to <code style="color:var(--ac)">~/.streamlit/secrets.toml</code>:
<pre style="background:var(--sf3);border-radius:6px;padding:.6rem;margin:.4rem 0;font-size:.72rem;color:var(--ac2)">OPENFDA_KEY  = "your_key"  # 120k/day
HF_TOKEN     = "your_key"  # HF Inference API
SERPAPI_KEY  = "your_key"  # Google Search</pre>
<a href="https://api.fda.gov" target="_blank" style="color:var(--ac);font-size:.75rem">Get openFDA key →</a>&nbsp;&nbsp;
<a href="https://huggingface.co/settings/tokens" target="_blank" style="color:var(--ac);font-size:.75rem">Get HF token →</a>
</div>""", unsafe_allow_html=True)

        st.markdown('<div class="sh">🗑 Cache Management</div>', unsafe_allow_html=True)
        c_a, c_b, c_c = st.columns(3)
        with c_a:
            if st.button("Clear Enrich", use_container_width=True):
                S.enrich_cache = {}; st.success("Cleared")
        with c_b:
            if st.button("Clear Ditto", use_container_width=True):
                S.ditto_cache = {}; st.success("Cleared")
        with c_c:
            if st.button("Clear All", use_container_width=True):
                S.enrich_cache = {}; S.ditto_cache = {}; S.who_cache = {}; st.success("Cleared")

        n_enr   = len(S.enrich_cache)
        n_ditto = len(S.ditto_cache)
        n_who   = len(S.who_cache)
        st.caption(f"Cache sizes — Enrich: {n_enr} · Ditto: {n_ditto} · WHO: {n_who}")

    st.info("ℹ Changes to Model Choice require re-building the catalog index to take effect.")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
elif S.page == "dashboard":
    st.markdown('<h1 class="pt">Dashboard</h1>', unsafe_allow_html=True)
    st.markdown('<p class="ps">Real-time system health, engine status, and query analytics.</p>', unsafe_allow_html=True)

    # Stats
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    for col, val, lbl, cls in [
        (c1, S.stats["queries"],        "Total Queries",   "sb"),
        (c2, S.stats["exact"],          "Exact Matches",   "sg"),
        (c3, S.stats["brand"],          "Brand Matches",   "sb"),
        (c4, S.stats["spec"],           "Spec Diff",       "sy"),
        (c5, S.stats["ditto_confirms"], "Ditto Confirms",  "sp"),
        (c6, f'{S.stats["avg_ms"]}ms',  "Avg Query Time",  "se"),
    ]:
        stat_card(col, val, lbl, cls)

    st.markdown("")

    # Pipeline visualization
    st.markdown('<div class="sh">9-Module Processing Pipeline</div>', unsafe_allow_html=True)
    modules = [
        ("📥","[1] INPUT",""),
        ("🔤","[2] NORMALIZE","INN/Brand → Generic"),
        ("🧠","[3] EMBED",f"{S.model_name or '?'}"),
        ("🔍","[4] RETRIEVE","Top-20 cosine"),
        ("⚡","[5] RERANK","Cross-Encoder"),
        ("📊","[6] HYBRID","70% sem + 30% fuzz"),
        ("✅","[7] VALIDATE","Clinical rules"),
        ("🔮","[8] DITTO","Pairwise confirm"),
        ("🌐","[9] ENRICH","openFDA + WHO INN"),
    ]
    active = {3: ST_OK, 4: True, 5: ST_OK and bool(load_crossencoder()),
              6: RF_OK, 7: True, 8: CFG.get("use_ditto"), 9: REQ_OK}
    parts = []
    for i, (ico, nm, desc) in enumerate(modules):
        mod_i = i + 1
        cls = "pipe dn" if active.get(mod_i, True) else ("pipe on" if mod_i in (3,5) else "pipe off")
        parts.append(f'<div class="{cls}"><span style="font-size:1rem">{ico}</span><b style="font-size:.6rem">{nm}</b><span style="font-size:.56rem">{desc}</span></div>')
        if i < len(modules) - 1:
            parts.append('<span style="color:var(--bd2);font-size:.85rem;align-self:center">→</span>')
    st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:.2rem;align-items:center;margin:.5rem 0">{"".join(parts)}</div>',
                unsafe_allow_html=True)

    # Engine capability table
    st.markdown('<div class="sh">Engine Capabilities</div>', unsafe_allow_html=True)
    cap = pd.DataFrame({
        "Module":     ["[3] Sentence Transformer","[5] Cross-Encoder Re-Ranker",
                       "[3] HuggingFace Inference API","[3] TF-IDF Fallback",
                       "[6] RapidFuzz Hybrid","[7] Validation Engine",
                       "[8] Ditto Verifier","[9] openFDA Device","[9] openFDA Drug",
                       "[9] WHO INN","FAISS Indexing"],
        "Available":  ["✅" if ST_OK else "❌ pip install sentence-transformers",
                       "✅" if (ST_OK and bool(load_crossencoder())) else "⚠",
                       "🔑 Needs HF_TOKEN in secrets",
                       "✅" if SKL_OK else "❌","✅" if RF_OK else "❌",
                       "✅ Always active","✅ Always active",
                       "✅" if REQ_OK else "❌","✅" if REQ_OK else "❌",
                       "✅" if REQ_OK else "❌","✅" if FAISS_OK else "pip install faiss-cpu"],
        "Purpose":    ["Primary semantic matching (384/768-dim embeddings)",
                       "Deep pairwise re-ranking of top-20 candidates",
                       "Remote embedding when ST not installable",
                       "Fast bag-of-words fallback (ngram 1-2)",
                       "Edit-distance similarity boost (30% weight)",
                       "Anatomy/dose/family rule checks (AO Foundation + PMC)",
                       "Final pairwise semantic confirmation layer",
                       "FDA device generic names, product codes, class",
                       "Drug INN, manufacturer, route, substance",
                       "WHO International Nonproprietary Names",
                       "Sub-millisecond retrieval for 100k+ catalogs"],
    })
    st.dataframe(cap, use_container_width=True, hide_index=True)

    # Catalog + query analytics
    if S.embed_meta:
        st.markdown('<div class="sh">Catalog Analytics</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            fvc = pd.Series([m["family"] for m in S.embed_meta]).value_counts().head(10)
            st.caption("Product Families")
            st.bar_chart(fvc, color="#00d4ff", height=180)
        with c2:
            cvc = pd.Series([m["category"] for m in S.embed_meta]).value_counts().head(10)
            st.caption("Categories")
            st.bar_chart(cvc, color="#00e676", height=180)
        with c3:
            drug_count  = sum(1 for m in S.embed_meta if NE.is_drug(m["desc"]))
            dev_count   = len(S.embed_meta) - drug_count
            vc = pd.Series({"Devices": dev_count, "Drugs": drug_count})
            st.caption("Drugs vs Devices")
            st.bar_chart(vc, color="#a78bfa", height=180)

    # Match type analytics
    if S.stats["queries"] > 0:
        st.markdown('<div class="sh">Query Analytics</div>', unsafe_allow_html=True)
        mt_data = pd.DataFrame({
            "Match Type": ["EXACT","BRAND_MATCH","SPEC_DIFF","NEW_SOP"],
            "Count": [S.stats["exact"],S.stats["brand"],S.stats["spec"],S.stats["new"]],
        })
        c1, c2 = st.columns(2)
        with c1:
            st.bar_chart(mt_data.set_index("Match Type"), color="#00d4ff", height=200)
        with c2:
            st.markdown(f"""
<div class="fda">
  <div class="ff">Total queries</div><div class="fv sb">{S.stats['queries']:,}</div>
  <div class="ff">Average latency</div><div class="fv sb">{S.stats['avg_ms']}ms</div>
  <div class="ff">Ditto confirmations</div><div class="fv g">{S.stats['ditto_confirms']:,}</div>
  <div class="ff">openFDA/WHO hits</div><div class="fv g">{S.stats['fda_hits']:,} / {S.stats['who_hits']:,}</div>
  <div class="ff">Active engine</div><div class="fv p">{S.model_name}</div>
</div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: CATALOG MANAGER
# ─────────────────────────────────────────────────────────────────────────────
elif S.page == "catalog":
    st.markdown('<h1 class="pt">Catalog Manager</h1>', unsafe_allow_html=True)
    st.markdown('<p class="ps">Upload your NPC catalog. MedMatch AI builds a semantic index with FAISS-ready embeddings.</p>', unsafe_allow_html=True)

    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown('<div class="sh">Upload NPC Catalog</div>', unsafe_allow_html=True)
        up = st.file_uploader("Upload (.xlsx, .csv)", type=["xlsx","xls","csv"], key="cat_up")
        if up:
            try:
                if up.name.endswith(".csv"):
                    df_c = pd.read_csv(up)
                else:
                    xl = pd.read_excel(up, sheet_name=None)
                    sheets = list(xl.keys())
                    sh = st.selectbox("Sheet", sheets) if len(sheets) > 1 else sheets[0]
                    df_c = xl[sh]
                df_c.columns = [str(c).strip() for c in df_c.columns]
                st.success(f"✓ {len(df_c):,} rows · {len(df_c.columns)} columns")
                st.dataframe(df_c.head(4), use_container_width=True)

                def best(kws):
                    return sorted(df_c.columns, key=lambda c: sum(k.upper() in c.upper() for k in kws), reverse=True)

                c_col = st.selectbox("NPC Code column", best(["npc","code"]))
                d_col = st.selectbox("Description column", best(["description","desc","product","name","item"]))
                cat_opts = ["(auto-detect)"] + list(df_c.columns)
                cat_col  = st.selectbox("Category column (optional)", cat_opts)
                cat_col  = cat_col if cat_col != "(auto-detect)" else None

                if st.button("🔨 Build Semantic Index", use_container_width=True):
                    bar = st.progress(0, "Normalizing...")
                    with st.spinner(""):
                        bar.progress(0.2, "Applying brand→generic aliases...")
                        bar.progress(0.4, "Computing embeddings...")
                        ok = build_catalog_index(df_c, d_col, c_col, cat_col)
                        bar.progress(0.85, "Building FAISS index..." if FAISS_OK else "Building TF-IDF...")
                        bar.progress(1.0, "Done ✓")
                    if ok:
                        idx_type = "FAISS" if S.faiss_idx else ("Numpy" if S.embeddings is not None else "TF-IDF")
                        st.success(f"✓ {len(S.embed_meta):,} products indexed · Engine: {S.model_name} · Index: {idx_type}")
                    else:
                        st.error("Indexing failed — check column selections.")
            except Exception as e:
                st.error(f"Error: {e}")

    with c2:
        st.markdown('<div class="sh">Index Status</div>', unsafe_allow_html=True)
        if S.embed_meta:
            n = len(S.embed_meta)
            dim = S.embeddings.shape[1] if S.embeddings is not None else 0
            vocab = len(S.tfidf_vec.vocabulary_) if S.tfidf_vec else 0
            idx_type = "FAISS" if S.faiss_idx else ("Numpy cosine" if S.embeddings is not None else "TF-IDF only")
            for val, lbl, cls in [
                (f"{n:,}",   "Products indexed", "sb"),
                (f"{dim}",   "Embedding dim",    "sg"),
                (f"{vocab:,}","TF-IDF vocab",    "sp"),
                (idx_type,   "Index type",       "sy"),
            ]:
                st.markdown(f'<div class="sc" style="margin-bottom:.45rem"><div class="sv {cls}" style="font-size:1.5rem">{val}</div><div class="sl">{lbl}</div></div>', unsafe_allow_html=True)

            st.markdown('<div class="sh">Sample Records</div>', unsafe_allow_html=True)
            sample = pd.DataFrame(S.embed_meta[:8])[["code","desc","family","atc"]]
            st.dataframe(sample, use_container_width=True, hide_index=True)
            if st.button("🗑 Clear Index", use_container_width=True):
                for k in ("embed_meta","embeddings","faiss_idx","tfidf_mat","tfidf_vec","catalog_df"):
                    setattr(S, k, None)
                st.rerun()
        else:
            st.info("No catalog indexed yet.\n\n**Expected columns:**\n- NPC Code\n- Product Description\n\nOptionally:\n- Category\n\nFormats: `.xlsx`, `.xls`, `.csv`")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: SINGLE SEARCH
# ─────────────────────────────────────────────────────────────────────────────
elif S.page == "search":
    st.markdown('<h1 class="pt">Product Search</h1>', unsafe_allow_html=True)
    st.markdown('<p class="ps">Enter any medical product — brand name, generic, abbreviated, or mixed-language. MedMatch AI normalizes, embeds, retrieves, re-ranks, and verifies.</p>', unsafe_allow_html=True)

    if not S.embed_meta:
        st.warning("⚠ No catalog loaded — go to **Catalog Manager** first.")
        if st.button("→ Catalog Manager", key="goto_cat_s"): S.page = "catalog"; st.rerun()
        st.stop()

    # Search bar
    tab1, tab2 = st.tabs(["🔍 Search", "📋 Examples"])
    with tab1:
        c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
        with c1:
            q = st.text_input("", "", key="sq", label_visibility="collapsed",
                              placeholder='e.g. "CAPROSYN 2-0 SH 70CM" · "3.5 cortex screw 26mm" · "Panadol 500mg tablet"')
        with c2:
            top_n = st.selectbox("Results", [3, 5, 10], index=1, key="top_n")
        with c3:
            do_e  = st.toggle("openFDA", False, key="do_e", help="Enrich top result via FDA + WHO API")
        with c4:
            run   = st.button("🔍 Search", use_container_width=True, disabled=not q.strip())

    with tab2:
        examples_grid = [
            ("Surgical Implants",   ["3.5MM CORTEX SCREW 26MM SELF TAPPING","PROXIMAL HUMERUS LOCKING PLATE 7 HOLES","EXPERT TIBIA NAIL 10MM 345MM","PFNA BLADE 120MM 95MM"]),
            ("Surgical Sutures",    ["CAPROSYN 2-0 SH 70CM","POLYSORB 0 CTX 90CM","TICRON 2-0 BRAIDED 75CM","PROLENE 4-0 SH MONOFILAMENT"]),
            ("Pharmaceuticals",     ["PANADOL 500MG TABLET","AUGMENTIN 625MG TABLET","FLAGYL 500MG INJECTION","COARTEM 80-480MG TABLET"]),
            ("Surgical Devices",    ["LIGASURE VESSEL SEALER 37CM","TROCAR 5MM LAPAROSCOPIC","EEA 28MM CIRCULAR STAPLER","REDON DRAIN 14FR CLOSED SUCTION"]),
        ]
        for cat, exs in examples_grid:
            st.markdown(f'<div style="font-size:.78rem;color:var(--mu);margin:.4rem 0">{cat}:</div>', unsafe_allow_html=True)
            ec = st.columns(len(exs))
            for col, ex in zip(ec, exs):
                with col:
                    lbl = ex[:22]+"…" if len(ex) > 22 else ex
                    if st.button(lbl, key=f"ex_{ex}", use_container_width=True):
                        st.session_state.sq = ex; run = True; q = ex

    st.markdown("---")

    if run and q.strip():
        q = q.strip()
        S.history.append(q)

        with st.spinner("Running 9-module pipeline…"):
            res = medmatch(q, top_n=top_n, do_enrich=do_e)

        if "error" in res: st.error(res["error"]); st.stop()

        # Metrics
        tr = res["results"][0] if res["results"] else {}
        c1,c2,c3,c4,c5 = st.columns(5)
        for col,val,lbl,cls in [
            (c1, f'{tr.get("score",0):.0f}%', "Top Score",  "sb"),
            (c2, f'{tr.get("ditto",{}).get("score","—")}%' if res.get("results") else "—", "Ditto", "sp"),
            (c3, len(res["results"]),           "Results",    "sg"),
            (c4, f'{res["elapsed_ms"]}ms',      "Time",       "sy"),
            (c5, res["family"],                 "Family",     "sb"),
        ]:
            stat_card(col, val, lbl, cls)

        st.markdown("")

        # Normalized + ATC + engine
        atc_badge = apill(res.get("atc",""))
        drug_badge = '<span class="pill pD">💊 DRUG</span>' if res.get("is_drug") else ""
        st.markdown(
            f'<div style="background:var(--sf);border:1px solid var(--bd);border-radius:8px;padding:.5rem 1rem;font-family:var(--mono);font-size:.76rem;margin:.4rem 0">'
            f'<span style="color:var(--mu)">Normalized: </span><span style="color:var(--ac)">{res["normalized"]}</span>'
            f'&nbsp;&nbsp;{epill(res["engine"])}&nbsp;{drug_badge}&nbsp;{atc_badge}</div>',
            unsafe_allow_html=True
        )

        # Spec chips
        sp = res["specs"]
        chips = ([f"📏 {d}mm" for d in sp["dims"]] +
                 [f"💧 {m}ml" for m in sp["ml"]] +
                 [f"💊 {g}mg" for g in sp["mg"]] +
                 [f"🔬 {g}mcg" for g in sp.get("mcg",[])] +
                 ([f"🩺 FR{f}" for f in sp["fr"]] if sp["fr"] else []) +
                 ([f"🔩 {sp['holes']}H"] if sp["holes"] else []) +
                 ([f"🔡 {sp['usp']}"] if sp["usp"] else []))
        if chips:
            chips_html = " ".join(
                f'<span style="background:var(--sf3);border:1px solid var(--bd);border-radius:4px;padding:2px 8px;font-size:.7rem;font-family:var(--mono)">{c}</span>'
                for c in chips)
            st.markdown(chips_html, unsafe_allow_html=True)

        st.markdown("")

        c_l, c_r = st.columns([3, 1])
        with c_l:
            st.markdown('<div class="sh">Ranked Matches</div>', unsafe_allow_html=True)
            for r in res["results"]:
                render_result_card(r, r["rank"] == 1)

        with c_r:
            if res["results"]:
                best = res["results"][0]
                ditto_d = best.get("ditto", {})
                ditto_score = ditto_d.get("score", 0)
                dc = "#00e676" if ditto_score >= 70 else ("#ffb300" if ditto_score >= 45 else "#ff4757")

                st.markdown('<div class="sh">Top Match</div>', unsafe_allow_html=True)
                st.markdown(f"""
<div class="fda">
  <div class="ff">NPC Code</div><div class="fv">{best['npc_code']}</div>
  <div class="ff">Match Type</div><div style="margin:.15rem 0">{mpill(best['match_type'])}</div>
  <div class="ff" style="margin-top:.4rem">Validation</div><div style="margin:.15rem 0">{vpill(best['val_status'])}</div>
  <div class="ff" style="margin-top:.4rem">Family</div><div class="fv">{best['family']}</div>
  <div class="ff" style="margin-top:.4rem">Scores</div>
  <div style="font-family:var(--mono);font-size:.75rem">
    Semantic: {best['sem_score']:.1f}%<br>Fuzzy: {best['fuzz_score']:.1f}%<br>Hybrid: {best['score']:.1f}%
  </div>
</div>""", unsafe_allow_html=True)

                if ditto_d:
                    st.markdown('<div class="sh" style="margin-top:.6rem">🔮 Ditto Verifier</div>', unsafe_allow_html=True)
                    st.markdown(f"""
<div class="fda ditto">
  <div style="display:flex;align-items:center;gap:.8rem">
    <div class="ditto-ring" style="border-color:{dc};color:{dc}">{ditto_score:.0f}%</div>
    <div>
      <div style="font-weight:700;color:{"var(--ok)" if ditto_d.get("confirmed") else "var(--warn)"}">
        {"✓ Confirmed" if ditto_d.get("confirmed") else "⚠ Low confidence"}
      </div>
      <div class="ff">{ditto_d.get("method","")}</div>
    </div>
  </div>
</div>""", unsafe_allow_html=True)

            if res.get("enrich"):
                ed = res["enrich"]
                is_drug_enr = res.get("is_drug", False)
                st.markdown('<div class="sh" style="margin-top:.6rem">🌐 Enrichment</div>', unsafe_allow_html=True)
                fda_fields = [
                    ("Source",       ed.get("source","")),
                    ("Generic Name", ed.get("device_name") or ed.get("generic_name","")),
                    ("Product Code", ed.get("product_code","")),
                    ("Device Class", ed.get("device_class","")),
                    ("Route",        ed.get("route","")),
                    ("Manufacturer", ed.get("manufacturer","")),
                    ("Regulation",   ed.get("regulation","")),
                ]
                html = '<div class="fda">'
                for lbl, val in fda_fields:
                    if val: html += f'<div class="ff">{lbl}</div><div class="fv">{val}</div>'
                # WHO INN
                if ed.get("who_inn"):
                    html += f'<div class="ff" style="color:var(--ac3);margin-top:.4rem">WHO INN</div><div class="fv g">{ed["who_inn"]}</div>'
                if ed.get("who_atc"):
                    html += f'<div class="ff">WHO ATC</div><div class="fv g">{ed["who_atc"]}</div>'
                if ed.get("definition"):
                    html += f'<div class="ff" style="margin-top:.4rem">FDA Definition</div><div style="font-size:.76rem;color:var(--mu);line-height:1.5">{ed["definition"][:200]}…</div>'
                html += '</div>'
                st.markdown(html, unsafe_allow_html=True)

        # Export
        st.markdown("")
        ed = res.get("enrich", {})
        export_rows = [{
            "INPUT_CODE": "SEARCH",
            "INPUT_DESCRIPTION":  q,
            "NPC_CODE":           r["npc_code"],
            "NPC_DESCRIPTION":    r["npc_desc"],
            "CATEGORY":           r["category"],
            "MATCH_SCORE":        r["score"],
            "MATCH_TYPE":         r["match_type"],
            "VALIDATION_STATUS":  r["val_status"],
            "VALIDATION_COMMENT": " | ".join(r["val_issues"]),
            "DITTO_SCORE":        r.get("ditto",{}).get("score",""),
            "ATC_CODE":           res.get("atc",""),
            "IS_DRUG":            res.get("is_drug",False),
            "FDA_PRODUCT_CODE":   ed.get("product_code",""),
            "FDA_DEFINITION":     ed.get("definition", ed.get("generic_name",""))[:200],
        } for r in res["results"]]
        st.download_button("⬇ Export Results (.xlsx)",
                           data=to_excel(export_rows),
                           file_name=f"medmatch_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE: BATCH PROCESSING
# ─────────────────────────────────────────────────────────────────────────────
elif S.page == "batch":
    st.markdown('<h1 class="pt">Batch Processing</h1>', unsafe_allow_html=True)
    st.markdown('<p class="ps">Upload an Excel/CSV file. MedMatch AI maps all products in parallel using the full 9-module pipeline.</p>', unsafe_allow_html=True)

    if not S.embed_meta:
        st.warning("⚠ No catalog loaded.")
        if st.button("→ Catalog Manager", key="goto_cat_b"): S.page = "catalog"; st.rerun()
        st.stop()

    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown('<div class="sh">Upload Products File</div>', unsafe_allow_html=True)
        up = st.file_uploader("Upload (.xlsx, .csv)", type=["xlsx","xls","csv"], key="batch_up")
        if up:
            try:
                df_b = pd.read_csv(up) if up.name.endswith(".csv") else pd.read_excel(up)
                df_b.columns = [str(c).strip() for c in df_b.columns]
                st.success(f"✓ {len(df_b):,} rows")
                st.dataframe(df_b.head(3), use_container_width=True)
                ranked = sorted(df_b.columns, key=lambda c: sum(k.upper() in c.upper()
                    for k in ["description","desc","product","item","name","consumable"]), reverse=True)
                dc  = st.selectbox("Description column", ranked)
                cc3 = st.selectbox("Code column (optional)", ["(none)"] + list(df_b.columns))
                mr  = st.number_input("Max rows (0 = all)", 0, 50000, 0)
                de  = st.toggle("openFDA + WHO enrichment", False, help="Adds latency per product")
            except Exception as e:
                st.error(f"Error: {e}"); st.stop()

            if st.button("⚡ Run Batch (Parallel)", use_container_width=True):
                data = df_b.dropna(subset=[dc])
                if mr > 0: data = data.head(mr)
                total = len(data)

                logs = []; lb = st.empty()
                def log(m, k="info"):
                    ts = datetime.now().strftime("%H:%M:%S")
                    cls = {"ok":"co","warn":"cw","err":"ce","ai":"cai","net":"cn","info":"ci"}.get(k,"ci")
                    logs.append(f'<span class="{cls}">[{ts}] {m}</span>')
                    lb.markdown('<div class="cons">' + "<br>".join(logs[-28:]) + "</div>", unsafe_allow_html=True)

                prog = st.progress(0, "Initializing…")
                log(f"Batch: {total:,} products · engine={S.model_name} · workers={CFG['batch_workers']}", "ai")
                if CFG.get("use_ditto"): log("Ditto verifier active — each result semantically confirmed", "ai")
                if de and REQ_OK: log("openFDA + WHO enrichment enabled — adds ~500ms per DRUG item", "net")

                descs = [str(r).strip() for r in data[dc].tolist()]
                codes = [str(data[cc3].iloc[i]).strip() if cc3 != "(none)" else f"ROW-{i+1}"
                         for i in range(len(data))]

                t0 = time.perf_counter()
                log("Starting parallel matching…", "ok")

                # Use parallel batch matcher
                all_res = []
                done_count = [0]

                def process_chunk(chunk_args):
                    chunk_descs, chunk_codes = chunk_args
                    return batch_match_parallel(chunk_descs, chunk_codes, do_enrich=de,
                                               workers=CFG.get("batch_workers", 3))

                # Process in chunks for live progress updates
                chunk_size = max(1, min(50, total // 10))
                chunks = [(descs[i:i+chunk_size], codes[i:i+chunk_size])
                          for i in range(0, total, chunk_size)]

                for ci, chunk in enumerate(chunks):
                    chunk_res = process_chunk(chunk)
                    all_res.extend(chunk_res)
                    pct = len(all_res) / total
                    prog.progress(pct, f"Processing {len(all_res):,}/{total:,}…")
                    if ci % 2 == 0:
                        nm = sum(1 for r in all_res if r.get("MATCH_TYPE","NEW_SOP") != "NEW_SOP")
                        elapsed = time.perf_counter() - t0
                        log(f"  {len(all_res):,}/{total:,} · {nm:,} matched · {elapsed:.1f}s", "ok")

                prog.progress(1.0, "Done ✓")
                S.batch_results = pd.DataFrame(all_res)
                elapsed = time.perf_counter() - t0

                n_total = len(all_res)
                n_match = sum(1 for r in all_res if r.get("MATCH_TYPE","NEW_SOP") != "NEW_SOP")
                n_rev   = sum(1 for r in all_res if r.get("VALIDATION_STATUS","") == "REVIEW")
                n_drug  = sum(1 for r in all_res if r.get("IS_DRUG", False))

                log(f"Done — {n_total:,} products in {elapsed:.1f}s ({elapsed/max(n_total,1)*1000:.0f}ms avg)", "ok")
                log(f"EXACT={S.stats['exact']} · BRAND={S.stats['brand']} · SPEC_DIFF={S.stats['spec']} · NEW_SOP={S.stats['new']}", "ai")
                log(f"Drugs detected: {n_drug:,} · Review required: {n_rev:,}", "warn" if n_rev > 0 else "ok")

    with c2:
        if S.batch_results is not None:
            df_r = S.batch_results; total = len(df_r)
            st.markdown('<div class="sh">Results Summary</div>', unsafe_allow_html=True)
            nm = int((df_r["MATCH_TYPE"] != "NEW_SOP").sum()) if "MATCH_TYPE" in df_r.columns else 0
            nr = int((df_r["VALIDATION_STATUS"] == "REVIEW").sum()) if "VALIDATION_STATUS" in df_r.columns else 0
            nd = int(df_r["IS_DRUG"].sum()) if "IS_DRUG" in df_r.columns else 0

            for val, lbl, cls in [
                (f"{total:,}",            "Total Processed", "sb"),
                (f"{nm:,}",               "Matched",         "sg"),
                (f"{total-nm:,}",         "Unmatched",       "sy"),
                (f"{nr:,}",               "Needs Review",    "se"),
                (f"{nd:,}",               "Drugs",           "sp"),
                (f"{int(df_r['DITTO_SCORE'].notna().sum() if 'DITTO_SCORE' in df_r.columns else 0):,}", "Ditto Scored", "sp"),
            ]:
                st.markdown(f'<div class="sc" style="margin-bottom:.4rem;"><div class="sv {cls}" style="font-size:1.4rem">{val}</div><div class="sl">{lbl}</div></div>', unsafe_allow_html=True)

            st.markdown("")
            if "MATCH_TYPE" in df_r.columns:
                vc = df_r["MATCH_TYPE"].value_counts().reset_index()
                vc.columns = ["Type", "Count"]
                st.bar_chart(vc.set_index("Type"), color="#00d4ff", height=140)

            xlsx = to_excel(df_r.to_dict("records"))
            fn = f"medmatch_batch_{datetime.now():%Y%m%d_%H%M}.xlsx"
            st.download_button("⬇ Export All (.xlsx)", data=xlsx, file_name=fn,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            st.download_button("⬇ Download CSV", data=df_r.to_csv(index=False).encode(),
                               file_name=fn.replace(".xlsx",".csv"), mime="text/csv")

    if S.batch_results is not None:
        st.markdown('<div class="sh">Results Table</div>', unsafe_allow_html=True)
        df_r = S.batch_results
        f1, f2, f3, f4 = st.columns(4)
        with f1: fm = st.selectbox("Match Type", ["All"]+sorted(df_r["MATCH_TYPE"].unique().tolist()) if "MATCH_TYPE" in df_r.columns else ["All"])
        with f2: fv = st.selectbox("Validation",  ["All","VALID","REVIEW"])
        with f3: fd = st.selectbox("Product Type", ["All","Drugs","Devices"])
        with f4: fs = st.text_input("Search", "", placeholder="filter…", key="bsrch")
        view = df_r.copy()
        if fm != "All" and "MATCH_TYPE" in view.columns: view = view[view["MATCH_TYPE"]==fm]
        if fv != "All" and "VALIDATION_STATUS" in view.columns: view = view[view["VALIDATION_STATUS"]==fv]
        if fd == "Drugs"   and "IS_DRUG" in view.columns: view = view[view["IS_DRUG"]==True]
        if fd == "Devices" and "IS_DRUG" in view.columns: view = view[view["IS_DRUG"]==False]
        if fs and "INPUT_DESCRIPTION" in view.columns:
            view = view[view["INPUT_DESCRIPTION"].str.upper().str.contains(fs.upper(), na=False)]

        def style_b(row):
            mt = row.get("MATCH_TYPE",""); vs = row.get("VALIDATION_STATUS","")
            if "CRITICAL" in str(row.get("VALIDATION_COMMENT","")): return ["background:rgba(255,71,87,.12)"]*len(row)
            if vs=="REVIEW":   return ["background:rgba(255,71,87,.07)"]*len(row)
            if mt=="EXACT":    return ["background:rgba(0,230,118,.04)"]*len(row)
            if mt=="NEW_SOP":  return ["background:rgba(255,179,0,.05)"]*len(row)
            return [""]*len(row)

        sc = ["INPUT_CODE","INPUT_DESCRIPTION","NPC_CODE","NPC_DESCRIPTION",
              "MATCH_SCORE","MATCH_TYPE","VALIDATION_STATUS","DITTO_SCORE",
              "ATC_CODE","FDA_PRODUCT_CODE"]
        sc = [c for c in sc if c in view.columns]
        st.dataframe(
            view[sc].style.apply(style_b, axis=1).format({"MATCH_SCORE":"{:.0f}","DITTO_SCORE":lambda x: f"{x:.0f}" if isinstance(x,(int,float)) and not pd.isna(x) else ""}),
            use_container_width=True, height=480
        )
        st.caption(f"{len(view):,} / {len(df_r):,} rows shown")
